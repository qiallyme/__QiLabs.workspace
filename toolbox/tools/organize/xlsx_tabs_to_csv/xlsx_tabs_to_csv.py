#!/usr/bin/env python3
# xlsx_tabs_to_csv.py
# Purpose: Interactively export every worksheet tab from one XLSX workbook into
# separate CSV files and create an index file in the output folder.

from pathlib import Path
import argparse
import csv
import re
import shutil
from openpyxl import load_workbook


def safe_filename(name: str) -> str:
    name = name.strip()
    name = re.sub(r'[<>:"/\\|?*]', "_", name)
    name = re.sub(r"\s+", "_", name)
    return name or "sheet"


def cell_value(value):
    if value is None:
        return ""
    return value


def ask_path(prompt: str, must_exist: bool = False, file_only: bool = False) -> Path:
    while True:
        raw = input(prompt).strip().strip('"').strip("'")

        if not raw:
            print("Please enter a path.")
            continue

        path = Path(raw).expanduser()

        if must_exist and not path.exists():
            print(f"Path not found: {path}")
            continue

        if file_only and not path.is_file():
            print(f"That is not a file: {path}")
            continue

        return path.resolve()


def ask_yes_no(prompt: str, default: bool = False) -> bool:
    suffix = " [Y/n]: " if default else " [y/N]: "

    while True:
        answer = input(prompt + suffix).strip().lower()

        if not answer:
            return default

        if answer in {"y", "yes"}:
            return True

        if answer in {"n", "no"}:
            return False

        print("Please answer y or n.")


def export_xlsx_to_csv(xlsx_path: Path, out_dir: Path, overwrite: bool = False) -> Path:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    if xlsx_path.suffix.lower() != ".xlsx":
        raise ValueError("Input file must be a .xlsx file")

    if out_dir.exists():
        if overwrite:
            shutil.rmtree(out_dir)
        else:
            raise FileExistsError(f"Output folder already exists: {out_dir}")

    out_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(
        filename=xlsx_path,
        read_only=True,
        data_only=True
    )

    index_rows = [
        ["sheet_name", "csv_filename", "rows_exported", "columns_exported"]
    ]

    used_filenames = set()

    for sheet in workbook.worksheets:
        base_name = safe_filename(sheet.title)
        csv_name = f"{base_name}.csv"

        counter = 2
        while csv_name.lower() in used_filenames:
            csv_name = f"{base_name}_{counter}.csv"
            counter += 1

        used_filenames.add(csv_name.lower())
        csv_path = out_dir / csv_name

        row_count = 0
        max_columns = 0

        with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)

            for row in sheet.iter_rows(values_only=True):
                cleaned_row = [cell_value(value) for value in row]

                if not any(str(value).strip() for value in cleaned_row):
                    continue

                writer.writerow(cleaned_row)
                row_count += 1
                max_columns = max(max_columns, len(cleaned_row))

        index_rows.append([
            sheet.title,
            csv_name,
            row_count,
            max_columns
        ])

    index_path = out_dir / "_index.csv"

    with index_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(index_rows)

    workbook.close()
    return out_dir


def interactive_mode():
    print()
    print("XLSX Tabs to CSV Exporter")
    print("-------------------------")

    xlsx_path = ask_path(
        "Enter the path to the XLSX file: ",
        must_exist=True,
        file_only=True
    )

    if xlsx_path.suffix.lower() != ".xlsx":
        raise ValueError("That file is not an .xlsx file.")

    default_out = xlsx_path.parent / f"{xlsx_path.stem}_csv_export"

    print()
    print(f"Default output folder:")
    print(default_out)

    use_default = ask_yes_no("Use this output folder?", default=True)

    if use_default:
        out_dir = default_out
    else:
        out_dir = ask_path("Enter the output folder path: ")

    overwrite = False

    if out_dir.exists():
        print()
        print(f"Output folder already exists:")
        print(out_dir)
        overwrite = ask_yes_no("Overwrite it?", default=False)

        if not overwrite:
            raise SystemExit("Cancelled. No files were changed.")

    print()
    print("Exporting workbook tabs...")
    export_folder = export_xlsx_to_csv(xlsx_path, out_dir, overwrite=overwrite)

    print()
    print("Done.")
    print(f"CSV export folder: {export_folder}")
    print(f"Index file: {export_folder / '_index.csv'}")


def main():
    parser = argparse.ArgumentParser(
        description="Export every tab in an XLSX workbook to separate CSV files."
    )

    parser.add_argument(
        "xlsx_file",
        nargs="?",
        help="Optional path to the .xlsx file"
    )

    parser.add_argument(
        "--out",
        help="Optional output folder path"
    )

    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite output folder if it already exists"
    )

    args = parser.parse_args()

    if not args.xlsx_file:
        interactive_mode()
        return

    xlsx_path = Path(args.xlsx_file).expanduser().resolve()

    if args.out:
        out_dir = Path(args.out).expanduser().resolve()
    else:
        out_dir = xlsx_path.parent / f"{xlsx_path.stem}_csv_export"

    export_folder = export_xlsx_to_csv(
        xlsx_path=xlsx_path,
        out_dir=out_dir,
        overwrite=args.overwrite
    )

    print("Done.")
    print(f"CSV export folder: {export_folder}")
    print(f"Index file: {export_folder / '_index.csv'}")


if __name__ == "__main__":
    main()